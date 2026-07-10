#!/usr/bin/env python3
"""Build Vendasta catalog spreadsheet with wholesale pricing."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Product Catalog"

# ── Styles ──────────────────────────────────────────────────────────────────
FONT_MAIN = "Arial"
HDR_BG  = "1F4E79"   # dark blue header
CAT_BG  = "2E75B6"   # section header
SUB_BG  = "BDD7EE"   # sub-section
ROW_ALT = "EBF3FB"   # alternating row
ADDON_BG= "F2F2F2"   # add-on rows

def hdr_font(size=11, bold=True, color="FFFFFF"):
    return Font(name=FONT_MAIN, size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="000000"):
    return Font(name=FONT_MAIN, size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

thin = Side(border_style="thin", color="BFBFBF")
border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

def set_row(ws, row, values, font=None, bg=None, bold=False, center_cols=None):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = font or body_font(bold=bold)
        if bg:
            c.fill = fill(bg)
        c.border = border_thin
        c.alignment = Alignment(
            horizontal="center" if (center_cols and col in center_cols) else "left",
            vertical="center", wrap_text=True
        )

# ── Columns ─────────────────────────────────────────────────────────────────
COLS = ["Category", "Product Name", "Edition / Tier", "Vendor", "Type",
        "Selling", "Active Accts", "SKU / Product ID",
        "Wholesale $/mo", "Retail $/mo", "Notes"]
COL_W = [22, 40, 20, 22, 12, 8, 10, 44, 14, 12, 35]

# Header row
r = 1
set_row(ws, r, COLS, font=hdr_font(11, True, "FFFFFF"), bg=HDR_BG, center_cols={1,3,5,6,7,8,9,10})
ws.row_dimensions[r].height = 20
r += 1

def cat_row(label):
    global r
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    c = ws.cell(r, 1, label)
    c.font = Font(name=FONT_MAIN, size=11, bold=True, color="FFFFFF")
    c.fill = fill(CAT_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = border_thin
    ws.row_dimensions[r].height = 18
    r += 1

def sub_row(label):
    global r
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    c = ws.cell(r, 1, label)
    c.font = Font(name=FONT_MAIN, size=10, bold=True, color="1F4E79")
    c.fill = fill(SUB_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = border_thin
    ws.row_dimensions[r].height = 15
    r += 1

alt = False
def data_row(cat, name, edition, vendor, typ, selling, accounts, sku, wholesale, retail, notes, is_addon=False):
    global r, alt
    bg = ADDON_BG if is_addon else (ROW_ALT if alt else "FFFFFF")
    alt = not alt
    vals = [cat, name, edition, vendor, typ,
            "✓" if selling else "–", accounts if accounts else "–",
            sku, wholesale, retail, notes]
    set_row(ws, r, vals, bg=bg, center_cols={1,3,5,6,7,9,10})
    # Format wholesale/retail as currency where numeric
    for col in (9, 10):
        cell = ws.cell(r, col)
        if isinstance(cell.value, (int, float)) and cell.value >= 0:
            cell.number_format = '"$"#,##0.00'
    ws.row_dimensions[r].height = 15
    r += 1

# ════════════════════════════════════════════════════════════════════════════
# 1. VENDASTA-NATIVE PRODUCTS
# ════════════════════════════════════════════════════════════════════════════
cat_row("VENDASTA-NATIVE PRODUCTS — Built and sold directly by Vendasta")

data_row("Vendasta","Vibe (Beta)","Free","Vendasta","Product",True,1,
    "MP-M56KKRF7Z3Z7NJFXRKDF67NXJLF8BN28:EDITION-VGFLLXC6",0.00,"","Free tier")
data_row("Vendasta","Vibe (Beta)","Standard","Vendasta","Product",True,1,
    "MP-M56KKRF7Z3Z7NJFXRKDF67NXJLF8BN28:EDITION-BKC3FL5Q",19.00,"","")
data_row("Vendasta","Vibe (Beta)","Pro","Vendasta","Product",True,0,
    "MP-M56KKRF7Z3Z7NJFXRKDF67NXJLF8BN28:EDITION-5HBC8RMQ","","","Wholesale TBD")
data_row("Vendasta","Vibe 10,000 credits/month","Add-on","Vendasta","Add-on",False,0,
    "A-F38XBKH52K","","","",True)
data_row("Vendasta","Vibe 25,000 credits/month","Add-on","Vendasta","Add-on",False,0,
    "A-G3X25FDFLM","","","",True)
data_row("Vendasta","Vibe 50,000 credits/month","Add-on","Vendasta","Add-on",False,0,
    "A-WFTVGDKLZN","","","",True)
data_row("Vendasta","Vibe 100,000 credits/month","Add-on","Vendasta","Add-on",False,0,
    "A-V6Z4BQ77JQ","","","",True)

data_row("Vendasta","CRM AI","Pro","Vendasta","Product",True,1,
    "MP-S8BSBT5P6MX4DM4PSZCFZLRMWFWDN8KZ:EDITION-4JS4G7R8",31.00,"","")
data_row("Vendasta","CRM AI","Standard","Vendasta","Product",True,0,
    "MP-S8BSBT5P6MX4DM4PSZCFZLRMWFWDN8KZ:EDITION-5W35ZXBN","","","Wholesale TBD")

data_row("Vendasta","Conversations AI","Standard","Vendasta","Product",True,0,
    "MP-NKC8VH78X2BKWJNPRSLH2TT3JWX83PD3:EDITION-8XGLKMNB",29.00,"","")
data_row("Vendasta","Conversations AI","Pro","Vendasta","Product",True,0,
    "MP-NKC8VH78X2BKWJNPRSLH2TT3JWX83PD3:EDITION-RBXDRVQP","","","Wholesale TBD")
data_row("Vendasta","Conversations AI","Premium","Vendasta","Product",True,4,
    "MP-NKC8VH78X2BKWJNPRSLH2TT3JWX83PD3:EDITION-F553FBQB",79.00,"","")

data_row("Vendasta","Reputation AI","Standard","Vendasta","Product",True,0,
    "RM:EDITION-F7JZ5TV8",0.00,"","Free tier — included in platform")
data_row("Vendasta","Reputation AI","Pro","Vendasta","Product",True,30,
    "RM",15.00,"","$15-17 seen in purchases; ~30 accounts")
data_row("Vendasta","Reputation AI","Premium","Vendasta","Product",True,0,
    "RM:EDITION-JFRPLQPN","","","Wholesale TBD")
data_row("Vendasta","Rapid Reviews Legacy","Add-on","Vendasta","Add-on",True,0,
    "RM:A-5QGW8G8VVG / RM:A-DVDGNKC4Q3",5.00,"","$5/mo per add-on",True)
data_row("Vendasta","Review Widget Pro","Add-on","Vendasta","Add-on",True,0,
    "A-5QGW8G8VVG","","","",True)

data_row("Vendasta","Social Marketing","Standard","Vendasta","Product",True,0,
    "SM:EDITION-FVGBNLVZ",4.00,"","")
data_row("Vendasta","Social Marketing","Pro","Vendasta","Product",True,66,
    "SM",15.00,"","~66 active accounts")

data_row("Vendasta","Local SEO","Standard","Vendasta","Product",True,0,
    "MS",0.00,"","Included / free tier")
data_row("Vendasta","Local SEO","Pro","Vendasta","Product",True,75,
    "MS:EDITION-CFH5CKHC",11.50,"","~75 active accounts")

sub_row("  Listings Add-ons (Vendasta)")
data_row("Vendasta","Citation Builder","Add-on to Local SEO","Vendasta","Add-on",True,1,
    "","","","Wholesale TBD",True)
data_row("Vendasta","Listing Sync Pro | USA (Monthly)","","Vendasta","Add-on",True,1,
    "LSP-0-US-M",17.00,"","",True)
data_row("Vendasta","Listing Sync Pro | USA (Legacy)","","Vendasta","Add-on",True,0,
    "","","","",True)
data_row("Vendasta","Listing Sync Pro | USA (Standard)","","Vendasta","Add-on",True,46,
    "LSP-1-US-M",25.00,"","~46 line-item appearances",True)
data_row("Vendasta","Listing Sync Pro | Australia (Monthly)","","Vendasta","Add-on",True,0,"","","","",True)
data_row("Vendasta","Listing Sync Pro | Australia Legacy","","Vendasta","Add-on",True,0,"","","","",True)
data_row("Vendasta","Listing Sync Pro | Canada (Monthly)","","Vendasta","Add-on",True,0,"","","","",True)
data_row("Vendasta","Listing Sync Pro | Canada","","Vendasta","Add-on",True,0,"","","","",True)
data_row("Vendasta","Listing Sync Pro | France (Monthly) Legacy","","Vendasta","Add-on",True,0,"","","","",True)
data_row("Vendasta","Listing Sync Pro | France Legacy","","Vendasta","Add-on",True,0,"","","","",True)
data_row("Vendasta","Listing Sync Pro | Germany (Monthly)","","Vendasta","Add-on",True,0,"","","","",True)
data_row("Vendasta","Listing Sync Pro | Germany Legacy","","Vendasta","Add-on",True,0,"","","","",True)
data_row("Vendasta","Listing Sync Pro | Ireland Legacy","","Vendasta","Add-on",True,0,"","","","",True)
data_row("Vendasta","Listing Sync Pro | Italy (Monthly)","","Vendasta","Add-on",True,0,"","","","",True)
data_row("Vendasta","Listing Sync Pro | Italy Legacy","","Vendasta","Add-on",True,0,"","","","",True)
data_row("Vendasta","Listing Sync Pro | UK (Monthly)","","Vendasta","Add-on",True,0,"","","","",True)
data_row("Vendasta","Listing Sync Pro | UK Legacy","","Vendasta","Add-on",True,0,"","","","",True)
data_row("Vendasta","Yext Listing Sync Pro (Monthly)","","Vendasta","Add-on",True,0,
    "A-8PHKXVRZFS","","","",True)
data_row("Vendasta","Yext Listing Sync Pro | USA","","Vendasta","Add-on",True,0,
    "A-WNW446NCNS","","","",True)

data_row("Vendasta","Campaigns Pro","","Vendasta","Product",True,1,
    "MP-ZGJ6V4QRP77WPMDKXS6VDRNX58Q42P7P",16.50,"","~23 purchase line items")

data_row("Vendasta","Customer Voice","Standard","Vendasta","Product",True,0,
    "MP-c4974d390a044c28aec31e421aa662b2:EDITION-TC8HJZNS",2.00,"","")
data_row("Vendasta","Customer Voice","Pro","Vendasta","Product",True,25,
    "MP-c4974d390a044c28aec31e421aa662b2",10.00,"","~25 active accounts")

sub_row("  Customer Voice SMS Add-ons")
sms_addons = [
    "Australia SMS 200 (Legacy)","Australia SMS 500","Australia SMS 1,000 (Legacy)",
    "Canada SMS 50","Canada SMS 100","Canada SMS 200","Canada SMS 500 (Legacy)",
    "Canada SMS 1,000 (Legacy)","UK SMS 200","UK SMS 500","UK SMS 1,000 (Legacy)",
    "US SMS 50","US SMS 100","US SMS 200 (Legacy)","US SMS 500","US SMS 1,000 (Legacy)",
]
for addon in sms_addons:
    data_row("Vendasta",addon,"","Vendasta","Add-on",True,0,"","","","",True)

data_row("Vendasta","WordPress Hosting","Standard","Vendasta","Product",True,0,
    "MP-ee4ea04e553a4b1780caf7aad7be07cd:EDITION-VFNL43ZF",2.00,"","")
data_row("Vendasta","WordPress Hosting","Pro","Vendasta","Product",True,72,
    "MP-ee4ea04e553a4b1780caf7aad7be07cd",12.00,"","~72 active accounts")
data_row("Vendasta","WordPress Hosting","Premium (Multisite)","Vendasta","Product",True,0,
    "MP-ee4ea04e553a4b1780caf7aad7be07cd:EDITION-6VFZW35G","","","Wholesale TBD")

data_row("Vendasta","Advertising Intelligence","","Vendasta","Product",True,52,
    "MP-94072e44d5364872b672d7ab4fc7a7e8",0.00,"","Free to partner; ~52 active accounts")
data_row("Vendasta","Advanced Reporting","Add-on to Adv. Intelligence","Vendasta","Add-on",True,0,
    "","","","Wholesale TBD",True)

data_row("Vendasta","Inbox Pro [Legacy]","","Vendasta","Product",True,0,
    "","","","Legacy product",)
data_row("Vendasta","AI-Assisted Web Chat [Legacy]","","Vendasta","Product",True,0,
    "","","","Legacy product")

data_row("Vendasta","Yesware","Premium","Vendasta","Product",True,2,
    "MP-NLMSDX765QS7QVWH44VCZJ2CFGKH58W2:EDITION-5TQDNSQV",35.00,"","")
data_row("Vendasta","Yesware Additional Seats","Add-on","Vendasta","Add-on",False,0,
    "A-8GNVGPBX2Z","","","",True)
data_row("Vendasta","CalendarHero Team Plan Seat","Add-on","Vendasta","Add-on",True,0,
    "","","","",True)

# ════════════════════════════════════════════════════════════════════════════
# 2. 0BYD CUSTOM SERVICES
# ════════════════════════════════════════════════════════════════════════════
cat_row("0BYD CUSTOM SERVICES — ImpactWorks / Rocket Local fulfillment (wholesale = $0)")

services_0byd = [
    ("AI Consulting Strategy Session", True, 0, ""),
    ("AI Essentials Automation", True, 0, ""),
    ("AI Growth Automation", True, 0, ""),
    ("Authority Website", True, 0, ""),
    ("Authority Website Activate", True, 0, ""),
    ("ADA Compliance (add-on to Authority Website)", True, 0, "Add-on"),
    ("ADA Compliance (add-on to Authority Website Activate)", True, 0, "Add-on"),
    ("Full Service SEO", True, 0, ""),
    ("Google Ads Management", False, 0, ""),
    ("Social Media Management", False, 0, ""),
    ("WS Web Services & Maintenance", True, 0, "MP-GSB376N5H3JFP52D6CTZFWMVPS3BL4RJ"),
    ("WU Web Updates", True, 0, "MP-7C5LDCT6T36ZDF55ZMKXWH5VV78RLDRZ"),
    ("Website Redesign & Development", True, 0, "MP-KQ5WCC4XNS5PQSDSNGHBP3DNMNPSSSNS"),
    ("ZAGG DTC Service", False, 0, "MP-QBF3T8NB22QBKKW5JFBRXK7S3CX57QRB"),
    ("RP Recolor Paint Collection Custom Web Application Development", False, 0, "MP-K2WGS37WMVSJR33QQLPRW3FWTCD3FJJS"),
]
for name, selling, accts, sku in services_0byd:
    data_row("0BYD Custom","Service: "+name,"","ImpactWorks / Rocket Local","Service",
             selling,accts,sku,0.00,"","Own fulfillment — no wholesale cost")

# ════════════════════════════════════════════════════════════════════════════
# 3. MARKETING SERVICES (Vendasta Fulfillment)
# ════════════════════════════════════════════════════════════════════════════
cat_row("MARKETING SERVICES — White-labeled fulfillment through Vendasta's agency arm")

sub_row("  Listings & GBP")
ms_listings = [
    ("Listing Claim Service", True, 0, "MP-935c959f76f34bb380aa86f49826d908"),
    ("Listing Claim: Medical Package", True, 0, ""),
    ("Additional Listing Claim or Re-verification (add-on)", True, 0, ""),
    ("Google Business Profile Verification & Claim", True, 1, "MP-FQ264D5N2822PT7ST6BVPFST6QHSKZC8"),
    ("Automated Google Search Ads (add-on to GBP Verification)", True, 0, ""),
    ("Google Business Profile Optimization", True, 0, ""),
    ("Edit & Update Listing Service (add-on to Listing Claim)", True, 0, ""),
    ("Local Listings Management Product", True, 0, ""),
]
wh_ls = {
    "Google Business Profile Verification & Claim": 60.00,
}
for name, selling, accts, sku in ms_listings:
    wh = wh_ls.get(name,"")
    data_row("Mktg Svcs",name,"","Vendasta","Service",selling,accts,sku,wh,"","")

sub_row("  Ads Campaigns")
ms_ads = [
    ("MatchCraft Managed Ads Campaign (req. Advertising Intelligence)", True, 0, "MP-WM65PB2J5PL6BCNLB5JJN58DDQWXZVPN"),
    ("  ↳ Boost (one-time) [Add-on]", True, 0, "A-C4X83LFNW7"),
    ("  ↳ Call Tracking [Add-on]", True, 0, "A-72TSXWWF24"),
    ("  ↳ Hourly Rate [Add-on]", True, 0, "A-WLZPJPCR7M"),
    ("  ↳ Landing Page [Add-on]", True, 0, "A-T8FLXQG7GP"),
    ("  ↳ Strategy Call: Monthly [Add-on]", True, 0, "A-XPG85MBF8X"),
    ("Amazon Sponsored Display Ads Campaign", False, 0, ""),
    ("  ↳ Spend Boost (one-time) [Add-on]", False, 0, ""),
    ("Digital Ads: Dynamic Automotive Legacy", True, 0, ""),
    ("  ↳ Additional Spend Plus [Add-on]", True, 0, ""),
    ("  ↳ Additional Spend [Add-on]", True, 0, ""),
    ("  ↳ Call Tracking [Add-on]", True, 0, ""),
    ("  ↳ Campaign [Add-on]", True, 0, ""),
    ("Monthly Additional Spend Legacy (add-on to Ads Robot)", True, 0, ""),
    ("Standard Display Banners", True, 0, "MP-2JB3LB2H8N77R325X84KBNP4XDTTMRGC"),
    ("  ↳ Additional Creative [Add-on]", True, 0, "A-VQJLVC4ZV5"),
    ("  ↳ Creative Edits [Add-on]", True, 0, "A-466MD7KMVJ"),
]
for name, selling, accts, sku in ms_ads:
    is_ao = name.startswith("  ↳")
    data_row("Mktg Svcs",name.strip(),"","Vendasta","Service",selling,accts,sku,"","","",is_ao)

sub_row("  Content & Blogging")
ms_content = [
    ("Blog Post: Monthly", True, 0, ""),
    ("  ↳ Additional Blog Post (Monthly) [Add-on]", True, 0, ""),
    ("  ↳ Custom Request (One-Time) [Add-on]", True, 0, ""),
    ("  ↳ Add 100 Words Monthly Legacy [Add-on]", True, 0, ""),
    ("  ↳ Add 100 Words One-time Legacy [Add-on]", True, 0, ""),
    ("Blog Post: One-Time", True, 0, ""),
    ("  ↳ Additional Blog Post [Add-on]", True, 0, ""),
]
for name, selling, accts, sku in ms_content:
    is_ao = name.startswith("  ↳")
    data_row("Mktg Svcs",name.strip(),"","Vendasta","Service",selling,accts,sku,"","","",is_ao)

sub_row("  Social Media")
ms_social = [
    ("Social Media Management Standard (req. Social Marketing)", True, 0, "MP-GKVMGQFB5HBKXNNKDX7SFVQPK7364FB7"),
    ("  ↳ Social Media Management Plus [Add-on]", True, 0, "A-MPSRNR886L"),
    ("Social Page Build", True, 0, "MP-912eba03d08242a3b9e449fe06255792"),
    ("  ↳ Additional Social Page Build [Add-on]", True, 0, "A-JP3VLR485M"),
    ("  ↳ Facebook Page Promotion Legacy [Add-on]", True, 0, ""),
]
for name, selling, accts, sku in ms_social:
    is_ao = name.startswith("  ↳")
    data_row("Mktg Svcs",name.strip(),"","Vendasta","Service",selling,accts,sku,"","","",is_ao)

sub_row("  Onboarding & Training")
ms_onboard = [
    ("Dashboard Walkthrough", True, 0, ""),
    ("  ↳ Additional 1-hour Training [Add-on]", True, 0, ""),
    ("AI Employee Setup", True, 2, ""),
    ("AI Workforce Optimization Plan", False, 0, ""),
]
for name, selling, accts, sku in ms_onboard:
    is_ao = name.startswith("  ↳")
    data_row("Mktg Svcs",name.strip(),"","Vendasta","Service",selling,accts,sku,"","","",is_ao)

sub_row("  Website")
ms_web = [
    ("Website Import Vetting", False, 0, "MP-1751871d98cf4e7f95d55c7d2b2ff293"),
    ("Website Support+ (req. WP Hosting)", False, 0, "MP-5FTSXB8XVSDRJW4RFQ84B8H3SGKG56JC"),
    ("  ↳ 30-Minute Call [Add-on]", False, 0, ""),
    ("  ↳ Additional Page [Add-on]", False, 0, ""),
    ("  ↳ Hourly Charge [Add-on]", False, 0, ""),
    ("Website Support Legacy (req. WP Hosting)", False, 0, ""),
    ("  ↳ 30-Minute Call Legacy [Add-on]", True, 0, "A-D63V6VCP85"),
    ("  ↳ Additional Page Legacy [Add-on]", True, 0, "A-7P6F4JPG67"),
    ("  ↳ Hourly Charge Legacy [Add-on]", True, 0, "A-PTBB4BP283"),
    ("  ↳ IDX Integration Legacy [Add-on]", True, 0, ""),
    ("  ↳ IDX Support Legacy [Add-on]", True, 0, ""),
    ("  ↳ Website Copy Legacy [Add-on]", True, 0, ""),
]
for name, selling, accts, sku in ms_web:
    is_ao = name.startswith("  ↳")
    data_row("Mktg Svcs",name.strip(),"","Vendasta","Service",selling,accts,sku,"","","",is_ao)

# ════════════════════════════════════════════════════════════════════════════
# 4. SM MARKETING INTERNATIONAL
# ════════════════════════════════════════════════════════════════════════════
cat_row("SM MARKETING INTERNATIONAL")
sm_int = [
    ("Alpha SEO - Full Service SEO", True, 0, "MP-XBRSXQSPSC4H4TSGMMMKHR3GQZMHL2LH", 175.00),
    ("Extra Geo (add-on)", True, 0, "", ""),
    ("Extra Topic (add-on)", True, 0, "", ""),
    ("Power Boost (add-on)", True, 0, "", ""),
    ("Max Power Boost (add-on)", True, 0, "", ""),
]
for name, selling, accts, sku, wh in sm_int:
    data_row("SM Marketing Intl",name,"","SM Marketing International","Product",selling,accts,sku,wh,"","")

# ════════════════════════════════════════════════════════════════════════════
# 5. BOOSTABILITY
# ════════════════════════════════════════════════════════════════════════════
cat_row("BOOSTABILITY")
data_row("Boostability","Local SEO Starter","","Boostability","Product",True,2,
    "MP-JP83X2HKPVP6NLVKGMRP7Z5N38HDKBHR:EDITION-GTP5PH5K",9.75,"","~2 active accounts")

# ════════════════════════════════════════════════════════════════════════════
# 6. GOOGLE
# ════════════════════════════════════════════════════════════════════════════
cat_row("GOOGLE")
google = [
    ("Google Workspace | Business Starter", True, 2, "MP-QBWGZ3KQ6SK6SJ6F64LWGFCL4PKQB7SR", 149.00),
    ("  ↳ Google Workspace Seat (add-on to Starter)", True, 3, "MP-94072e44d5364872b672d7ab4fc7a7e8:A-3QKQHBS3R6", 9.00),
    ("Google Workspace | Business Standard", True, 0, "", ""),
    ("  ↳ Google Workspace Seat (add-on to Standard)", True, 0, "", ""),
    ("Google Workspace | Enterprise Plus", True, 0, "", ""),
    ("  ↳ Google Workspace Seat (add-on to Enterprise)", True, 0, "", ""),
    ("Google Workspace Transfer", False, 0, "", ""),
]
for name, selling, accts, sku, wh in google:
    is_ao = name.startswith("  ↳")
    data_row("Google",name.strip(),"","Google","Product",selling,accts,sku,wh,"","",is_ao)

# ════════════════════════════════════════════════════════════════════════════
# 7. GODADDY
# ════════════════════════════════════════════════════════════════════════════
cat_row("GODADDY")
data_row("GoDaddy","GoDaddy Domains","","GoDaddy","Product",True,2,
    "MP-4TMLZSQ5FMJQX5T75TPC43FQBWD2VXLB",12.00,"","~4 purchase appearances")
data_row("GoDaddy","Domains","","GoDaddy","Product",True,0,"","","","")

# ════════════════════════════════════════════════════════════════════════════
# 8. SHOPTOIT (Legacy)
# ════════════════════════════════════════════════════════════════════════════
cat_row("SHOPTOIT (Legacy)")
shoptoit = [
    "Google Ads for Agencies Legacy",
    "Advertising Budget (add-on) Legacy",
    "Google Display Ads for Small Business Legacy",
    "One Time Budget Addition (add-on) Legacy",
    "Google Search Ads for Small Businesses Legacy",
    "One Time Budget Addition (add-on) Legacy (Search)",
    "QuickAds Legacy",
]
for name in shoptoit:
    is_ao = "(add-on)" in name
    data_row("Shoptoit",name,"","Shoptoit","Product",True,0,"","","","Legacy",is_ao)

# ════════════════════════════════════════════════════════════════════════════
# 9. SITEGLUE
# ════════════════════════════════════════════════════════════════════════════
cat_row("SITEGLUE")
data_row("SiteGlue","Custom AI Development (add-on to SiteGlue AI Starter)","","SiteGlue","Add-on",True,0,
    "A-D42DTCB3WC","","","")

# ════════════════════════════════════════════════════════════════════════════
# 10. ONLINE DESIGN CLUB
# ════════════════════════════════════════════════════════════════════════════
cat_row("ONLINE DESIGN CLUB")
odc = [
    ("Business Card Design", False),
    ("Additional Business Card Design (add-on)", True),
    ("Corporate Identity Design", False),
    ("Custom Brochure Design", False),
    ("Custom Graphic Design Project", False),
    ("Custom Label Design", False),
    ("Custom Logo Design", False),
    ("Custom Marketing Flyer Design", False),
    ("Custom Mascot Design", False),
    ("Social Media Profile Design", True),
]
for name, selling in odc:
    is_ao = "(add-on)" in name
    data_row("Online Design Club",name,"","Online Design Club","Service",selling,0,"","","","",is_ao)

# ════════════════════════════════════════════════════════════════════════════
# 11. MAGNFI
# ════════════════════════════════════════════════════════════════════════════
cat_row("MAGNFI")
data_row("Magnfi","Magnfi Video White Label Plan","","Magnfi","Product",False,0,"","","","")

# ════════════════════════════════════════════════════════════════════════════
# 12. ADCELLERANT (Suspended)
# ════════════════════════════════════════════════════════════════════════════
cat_row("ADCELLERANT — z-suspended")
data_row("AdCellerant","z-suspended - Targeted Display CA & AU Only Legacy","","AdCellerant","Product",True,0,"","","","Suspended")
data_row("AdCellerant","Targeted Display Additional Spend (add-on) Legacy","","AdCellerant","Add-on",True,0,"","","","Suspended",True)

# ════════════════════════════════════════════════════════════════════════════
# 13. ACTIVECAMPAIGN (Legacy)
# ════════════════════════════════════════════════════════════════════════════
cat_row("ACTIVECAMPAIGN (Legacy)")
data_row("ActiveCampaign","ActiveCampaign Legacy","","ActiveCampaign","Product",True,0,"","","","Legacy")

# ════════════════════════════════════════════════════════════════════════════
# 14. LOCALADS (Legacy)
# ════════════════════════════════════════════════════════════════════════════
cat_row("LOCALADS (Legacy)")
data_row("LocalAds","Marketing Insights Legacy","","LocalAds","Product",True,0,"","","","Legacy")
data_row("LocalAds","Audience Capture Legacy (add-on)","","LocalAds","Add-on",False,0,"","","","Legacy",True)
data_row("LocalAds","Customer List Marketing Legacy (add-on)","","LocalAds","Add-on",False,0,"","","","Legacy",True)
data_row("LocalAds","Premium Marketing Insights Legacy (add-on)","","LocalAds","Add-on",False,0,"","","","Legacy",True)

# ════════════════════════════════════════════════════════════════════════════
# 15. HYPERSIGN (Legacy)
# ════════════════════════════════════════════════════════════════════════════
cat_row("HYPERSIGN (Legacy)")
data_row("Hypersign","Hypersign Digital Signage Software | Pro Legacy","","Hypersign","Product",False,0,"","","","Legacy")
data_row("Hypersign","Digital Signage Media Player Package Legacy (add-on)","","Hypersign","Add-on",False,0,"","","","Legacy",True)
data_row("Hypersign","Hypersign Pro Multiple Licenses Legacy (add-on)","","Hypersign","Add-on",False,0,"","","","Legacy",True)

# ════════════════════════════════════════════════════════════════════════════
# 16. SO CONNECT (Legacy)
# ════════════════════════════════════════════════════════════════════════════
cat_row("SO CONNECT (Legacy)")
data_row("So Connect","Website Optimizer Legacy","","So Connect","Product",True,0,"","","","Legacy")

# ════════════════════════════════════════════════════════════════════════════
# Column widths
# ════════════════════════════════════════════════════════════════════════════
for i, w in enumerate(COL_W, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze top row + header
ws.freeze_panes = "A2"

# ════════════════════════════════════════════════════════════════════════════
# Sheet 2: Wholesale Pricing Reference
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Wholesale Reference")
ws2.freeze_panes = "A2"

ref_cols = ["SKU / Product ID", "Product Name (inferred)", "Wholesale $/mo", "Active Count (purchase lines)", "Notes"]
ref_col_w = [46, 42, 14, 18, 40]

r2 = 1
set_row(ws2, r2, ref_cols, font=hdr_font(11, True, "FFFFFF"), bg=HDR_BG)
ws2.row_dimensions[r2].height = 20
r2 += 1

wholesale_data = [
    ("SM", "Social Marketing Pro", 15.00, 275, "Seen $15-17/mo; ~275 billing records"),
    ("SM:EDITION-FVGBNLVZ", "Social Marketing Standard", 4.00, 2260, ""),
    ("MS", "Local SEO Standard", 0.00, 3165, "Appears free (included in platform sub)"),
    ("MS:EDITION-CFH5CKHC", "Local SEO Pro", 11.50, 80, ""),
    ("RM:EDITION-F7JZ5TV8", "Reputation AI Standard", 0.00, 910, "Free tier"),
    ("RM:EDITION-JFRPLQPN", "Reputation AI Premium", None, 0, "No purchase records yet"),
    ("RM", "Reputation AI Pro", 15.00, 32, "$15-17/mo seen"),
    ("RM-TRIAL", "Reputation AI Trial", 0.00, 38, "Trial — $0"),
    ("RM:A-5QGW8G8VVG", "Rapid Reviews Add-on", 5.00, 10, ""),
    ("RM:A-DVDGNKC4Q3", "Rapid Reviews Add-on (variant)", 5.00, 9, ""),
    ("MP-94072e44d5364872b672d7ab4fc7a7e8", "Advertising Intelligence", 0.00, 2694, "Free to partner"),
    ("MP-94072e44d5364872b672d7ab4fc7a7e8:A-3QKQHBS3R6", "Google Workspace Seat (Starter add-on)", 9.00, 20, "$9-9.50/mo"),
    ("MP-c4974d390a044c28aec31e421aa662b2:EDITION-TC8HJZNS", "Customer Voice Standard", 2.00, 738, ""),
    ("MP-c4974d390a044c28aec31e421aa662b2", "Customer Voice Pro", 10.00, 33, ""),
    ("MP-ee4ea04e553a4b1780caf7aad7be07cd:EDITION-VFNL43ZF", "WordPress Hosting Standard", 2.00, 3101, ""),
    ("MP-ee4ea04e553a4b1780caf7aad7be07cd", "WordPress Hosting Pro", 12.00, 1071, ""),
    ("MP-NKC8VH78X2BKWJNPRSLH2TT3JWX83PD3:EDITION-F553FBQB", "Conversations AI Premium", 79.00, 29, ""),
    ("MP-NKC8VH78X2BKWJNPRSLH2TT3JWX83PD3:EDITION-8XGLKMNB", "Conversations AI Standard", 29.00, 1, ""),
    ("MP-M56KKRF7Z3Z7NJFXRKDF67NXJLF8BN28:EDITION-BKC3FL5Q", "Vibe Standard", 19.00, 1, ""),
    ("MP-M56KKRF7Z3Z7NJFXRKDF67NXJLF8BN28:EDITION-VGFLLXC6", "Vibe Free", 0.00, 1, "Free tier"),
    ("MP-S8BSBT5P6MX4DM4PSZCFZLRMWFWDN8KZ:EDITION-4JS4G7R8", "CRM AI Pro", 31.00, 3, ""),
    ("MP-ZGJ6V4QRP77WPMDKXS6VDRNX58Q42P7P", "Campaigns Pro", 16.50, 23, ""),
    ("MP-NLMSDX765QS7QVWH44VCZJ2CFGKH58W2:EDITION-5TQDNSQV", "Yesware Premium", 35.00, 14, ""),
    ("MP-QBWGZ3KQ6SK6SJ6F64LWGFCL4PKQB7SR", "Google Workspace Business Starter", 149.00, 12, ""),
    ("MP-4TMLZSQ5FMJQX5T75TPC43FQBWD2VXLB", "GoDaddy Domains", 12.00, 4, ""),
    ("MP-FQ264D5N2822PT7ST6BVPFST6QHSKZC8", "GBP Verification & Claim", 60.00, 1, ""),
    ("MP-XBRSXQSPSC4H4TSGMMMKHR3GQZMHL2LH", "Alpha SEO - Full Service SEO", 175.00, 2, ""),
    ("MP-JP83X2HKPVP6NLVKGMRP7Z5N38HDKBHR:EDITION-GTP5PH5K", "Boostability Local SEO Starter", 9.75, 47, ""),
    ("LSP-0-US-M", "Listing Sync Pro USA (Starter)", 17.00, 35, "$17-18/mo"),
    ("LSP-1-US-M", "Listing Sync Pro USA (Standard)", 25.00, 46, "$25-30/mo"),
    ("ST", "Snapshot Report (est.)", 2.00, 225, "Short code; product TBD"),
    ("LD", "Legacy Local Digital (est.)", 50.00, 29, "Short code; product TBD — $50-53/mo"),
    ("MP-H58MV5X2ZXLLHCQ727JBVD8QF4T8XM5P", "Unknown product", 29.00, 38, "~1 account; verify in portal"),
    ("MP-DKT6XHPM6NCCDNK2TPDPVD3PG3V7ZHWP", "Unknown product", 39.00, 21, "Verify in portal"),
    ("MP-6GLFKDNCV3LTHTRZ6DDXLJN7MLVBBWGW", "Unknown product", 50.00, 11, "Verify in portal"),
    ("MP-JX7KQ83QRJB43SSFXKDG4STJZF7ZDB32", "Unknown product", 29.00, 6, "Verify in portal"),
    ("MP-G222MCLJMR86JCFZWDDVZDXZ7NXZWNKR", "Unknown product", 99.00, 7, "Verify in portal"),
    ("MP-JFW4X5KMBK44GT4SBDZL72KVP2DPZQZS", "Unknown product", 100.00, 10, "Verify in portal"),
    ("MP-XBHPSLDBHZ8Q8F57P43DPKL6SPHKHHMS", "Unknown product", 64.75, 5, "Verify in portal"),
    ("MP-K48XKQV7DXLGQZ67ZH4MNZDM2FQCK8LM", "Unknown product", 199.00, 2, "Verify in portal"),
    ("vbp2_growth_subscription", "Vendasta Platform — Growth Tier", 575.00, 49, "PARTNER subscription (not client product)"),
    ("vbp_basic_subscription", "Vendasta Platform — Basic Tier", 275.00, 31, "PARTNER subscription (not client product)"),
    ("sa-additional-market-subscription", "Additional Market Subscription", 250.00, 49, "Platform fee for additional market"),
    ("sa-premium-reports", "Premium Reports", 249.00, 7, ""),
    ("snapshotreportrefresh", "Snapshot Report Refresh", 0.00, 27, "Free add-on"),
]

alt2 = False
for row_data in wholesale_data:
    bg = ROW_ALT if alt2 else "FFFFFF"
    alt2 = not alt2
    vals = list(row_data)
    # Replace None with empty string
    vals = [v if v is not None else "" for v in vals]
    set_row(ws2, r2, vals, bg=bg, center_cols={3,4})
    if isinstance(vals[2], float):
        ws2.cell(r2, 3).number_format = '"$"#,##0.00'
    ws2.row_dimensions[r2].height = 14
    r2 += 1

for i, w in enumerate(ref_col_w, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════════════════════
out = "/Users/dantecrescenzi/claudeclaw/store/vendasta-product-catalog.xlsx"
wb.save(out)
print(f"Saved: {out}")
